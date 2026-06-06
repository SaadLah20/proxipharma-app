#!/usr/bin/env python3
"""
Convertit le fichier Excel officine (Article, Ppv, Pph, TVA) en CSV importable.

Ne garde que les médicaments : TVA = 0.
Dédoublonne par nom (insensible à la casse).

Usage :
  python scripts/convert-medicaments-xlsx.py "c:\\Users\\pc\\Downloads\\Base de données médicaments (1).xlsx"
  python scripts/convert-medicaments-xlsx.py --xlsx path/to/file.xlsx --out scripts/medicaments_officine.csv
"""

from __future__ import annotations

import argparse
import csv
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Installez openpyxl : pip install openpyxl", file=sys.stderr)
    raise SystemExit(1)

SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_OUT = SCRIPT_DIR / "medicaments_officine.csv"


def parse_price(value: object) -> str:
    if value is None:
        return ""
    s = str(value).strip().replace("\u00a0", " ").replace(" ", "").replace(",", ".")
    if not s:
        return ""
    try:
        n = float(s)
    except ValueError:
        return ""
    if n <= 0:
        return ""
    return f"{n:.2f}".rstrip("0").rstrip(".") if "." in f"{n:.2f}" else f"{n:.2f}"


def is_medicament_tva(tva: object) -> bool:
    if tva is None:
        return False
    if isinstance(tva, (int, float)):
        return tva == 0
    s = str(tva).strip().replace(",", ".")
    try:
        return float(s) == 0
    except ValueError:
        return s == "0"


def convert(xlsx_path: Path, out_path: Path) -> dict[str, int]:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    seen: set[str] = set()
    rows_out: list[dict[str, str]] = []
    stats = {"total_rows": 0, "medicament_rows": 0, "skipped_dup": 0, "skipped_bad": 0}

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        if not row or not row[0]:
            continue
        stats["total_rows"] += 1
        if not is_medicament_tva(row[3] if len(row) > 3 else None):
            continue
        stats["medicament_rows"] += 1
        name = re.sub(r"\s+", " ", str(row[0]).strip())
        if not name:
            stats["skipped_bad"] += 1
            continue
        key = name.casefold()
        if key in seen:
            stats["skipped_dup"] += 1
            continue
        seen.add(key)
        price_ppv = parse_price(row[1] if len(row) > 1 else None)
        price_pph = parse_price(row[2] if len(row) > 2 else None)
        if not price_ppv or not price_pph:
            stats["skipped_bad"] += 1
            continue
        rows_out.append({"name": name, "price_ppv": price_ppv, "price_pph": price_pph})

    wb.close()

    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["name", "price_ppv", "price_pph"])
        writer.writeheader()
        writer.writerows(rows_out)

    stats["written"] = len(rows_out)
    return stats


def main() -> int:
    parser = argparse.ArgumentParser(description="Excel médicaments → CSV (TVA=0)")
    parser.add_argument("xlsx", nargs="?", type=Path, help="Fichier .xlsx source")
    parser.add_argument("--xlsx", type=Path, dest="xlsx_flag", help="Chemin .xlsx")
    parser.add_argument("--out", type=Path, default=DEFAULT_OUT, help="CSV de sortie")
    args = parser.parse_args()
    xlsx = args.xlsx_flag or args.xlsx
    if not xlsx or not xlsx.is_file():
        print("Fichier xlsx introuvable.", file=sys.stderr)
        return 1
    stats = convert(xlsx, args.out)
    print(f"Source      : {xlsx}")
    print(f"CSV         : {args.out}")
    print(f"Lignes lues : {stats['total_rows']}")
    print(f"TVA=0       : {stats['medicament_rows']}")
    print(f"Dédoublonnés: {stats['skipped_dup']}")
    print(f"Rejetés     : {stats['skipped_bad']}")
    print(f"Écrits      : {stats['written']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
